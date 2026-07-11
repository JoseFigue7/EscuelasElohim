import io
import os

from django.conf import settings
from django.core.files.base import ContentFile
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.db.models import Q, Count
from decimal import Decimal

from .models import (
    Curso, Promocion, Tema, Material, Inscripcion, 
    Asistencia, Pregunta, Examen, RespuestaExamen, RecuperacionExamen,
    CalificacionExamen, PromedioPromocion, Diploma
)
from .serializers import (
    CursoSerializer, PromocionSerializer, TemaSerializer, TemaListSerializer,
    MaterialSerializer, InscripcionSerializer, AsistenciaSerializer,
    PreguntaSerializer, PreguntaDetailSerializer, ExamenSerializer, ExamenListSerializer,
    RespuestaExamenSerializer, RecuperacionExamenSerializer, RecuperacionExamenAlumnoSerializer,
    RecuperacionExamenBulkCreateSerializer,
    CalificacionExamenSerializer, CalificacionExamenDetalleSerializer, PromedioPromocionSerializer, DiplomaSerializer,
    RespuestaRevisionSerializer
)


class CursoViewSet(viewsets.ModelViewSet):
    queryset = Curso.objects.all()
    serializer_class = CursoSerializer
    permission_classes = [IsAuthenticated]


class PromocionViewSet(viewsets.ModelViewSet):
    queryset = Promocion.objects.select_related('curso', 'docente').prefetch_related('docentes').all()
    serializer_class = PromocionSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        
        # Alumnos solo ven promociones donde están inscritos
        if user.es_alumno:
            inscripciones = Inscripcion.objects.filter(
                alumno=user, activa=True
            ).values_list('promocion_id', flat=True)
            queryset = queryset.filter(id__in=inscripciones, activa=True)
        
        # Docentes ven promociones asignadas; administradores ven todas
        elif user.tipo == 'docente':
            queryset = queryset.filter(
                Q(docente=user) | Q(docentes=user)
            ).distinct()
        
        return queryset
    
    def perform_create(self, serializer):
        """Asignar automáticamente el docente al usuario actual si es docente o admin"""
        user = self.request.user
        if user.es_docente or user.is_superuser:
            promocion = serializer.save(docente=user)
            if not promocion.docentes.filter(pk=user.pk).exists():
                promocion.docentes.add(user)
        else:
            serializer.save()

    @action(detail=True, methods=['get'])
    def exportar_notas(self, request, pk=None):
        """Descarga un Excel con notas por alumno (filas) y tema (columnas). Solo docente/admin."""
        user = request.user
        if not (user.es_docente or user.is_superuser):
            return Response(
                {'error': 'Solo docentes y administradores pueden descargar las notas'},
                status=status.HTTP_403_FORBIDDEN,
            )

        promocion = self.get_object()
        temas = list(
            Tema.objects.filter(curso=promocion.curso)
            .select_related('examen')
            .order_by('numero_tema')
        )
        inscripciones = list(
            Inscripcion.objects.filter(promocion=promocion, activa=True)
            .select_related('alumno')
            .order_by('alumno__last_name', 'alumno__first_name', 'alumno__username')
        )

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from django.http import HttpResponse
        from urllib.parse import quote

        wb = Workbook()
        ws = wb.active
        ws.title = 'Notas'

        headers = ['Alumno']
        for tema in temas:
            headers.append(f'T{tema.numero_tema} - {tema.titulo}')
        headers.extend(['Promedio', 'Estado'])
        ws.append(headers)

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='1F4E79')
        thin = Border(
            left=Side(style='thin', color='D0D7DE'),
            right=Side(style='thin', color='D0D7DE'),
            top=Side(style='thin', color='D0D7DE'),
            bottom=Side(style='thin', color='D0D7DE'),
        )
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin

        for inscripcion in inscripciones:
            alumno = inscripcion.alumno
            nombre = (alumno.get_full_name() or '').strip() or alumno.username
            row = [nombre]
            porcentajes = []

            for tema in temas:
                try:
                    examen = tema.examen
                except Examen.DoesNotExist:
                    row.append(None)
                    continue
                calificacion = CalificacionExamen.get_calificacion_efectiva(examen, inscripcion)
                if calificacion is None:
                    row.append(None)
                else:
                    porcentaje = round(float(calificacion.porcentaje), 2)
                    row.append(porcentaje)
                    porcentajes.append(porcentaje)

            if porcentajes:
                promedio = round(sum(porcentajes) / len(porcentajes), 2)
                estado = 'Aprobado' if promedio >= 80 else 'Reprobado'
            else:
                promedio = None
                estado = ''

            row.extend([promedio, estado])
            ws.append(row)

            excel_row = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.border = thin
                if col_idx > 1:
                    cell.alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 32
        for col_idx in range(2, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
        ws.row_dimensions[1].height = 36
        ws.freeze_panes = 'B2'

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        base_name = promocion.nombre or f'promocion_{promocion.id}'
        safe_name = ''.join(ch if ch.isalnum() or ch in ('-', '_') else '_' for ch in base_name)
        safe_name = safe_name.strip('_') or f'promocion_{promocion.id}'
        filename = f'Notas_{safe_name}.xlsx'
        filename_encoded = quote(filename, safe='')

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename_encoded}'
        )
        return response


class TemaViewSet(viewsets.ModelViewSet):
    queryset = Tema.objects.select_related('curso').prefetch_related('materiales').all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'list':
            return TemaListSerializer
        return TemaSerializer
    
    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        
        # Alumnos solo ven temas de cursos de promociones donde están inscritos
        if user.es_alumno:
            inscripciones = Inscripcion.objects.filter(
                alumno=user, activa=True
            ).select_related('promocion__curso')
            curso_ids = [insc.promocion.curso_id for insc in inscripciones]
            queryset = queryset.filter(
                curso_id__in=curso_ids,
                visible_para_estudiante=True,
            )
        
        # Filtrar por promoción: obtener el curso de la promoción
        promocion_id = self.request.query_params.get('promocion')
        if promocion_id:
            try:
                promocion = Promocion.objects.select_related('curso').get(id=promocion_id)
                queryset = queryset.filter(curso_id=promocion.curso_id)
            except Promocion.DoesNotExist:
                queryset = queryset.none()
        
        # Filtrar por curso si se proporciona directamente
        curso_id = self.request.query_params.get('curso')
        if curso_id:
            queryset = queryset.filter(curso_id=curso_id)
        
        return queryset


class MaterialViewSet(viewsets.ModelViewSet):
    queryset = Material.objects.select_related('tema', 'tema__curso').all()
    serializer_class = MaterialSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        
        # Alumnos solo ven materiales de temas de cursos donde están inscritos
        if user.es_alumno:
            inscripciones = Inscripcion.objects.filter(
                alumno=user, activa=True
            ).select_related('promocion__curso')
            curso_ids = [insc.promocion.curso_id for insc in inscripciones]
            queryset = queryset.filter(
                tema__curso_id__in=curso_ids,
                tema__visible_para_estudiante=True,
            )
        
        # Filtrar por tema si se proporciona
        tema_id = self.request.query_params.get('tema')
        if tema_id:
            queryset = queryset.filter(tema_id=tema_id)
        
        return queryset
    
    def retrieve(self, request, *args, **kwargs):
        """Sobrescribir retrieve para servir el archivo con el tipo de contenido correcto"""
        instance = self.get_object()
        
        # Si se solicita descargar el archivo (query param download=true)
        if request.query_params.get('download') == 'true' and instance.archivo:
            from django.http import FileResponse
            import os
            from mimetypes import guess_type
            from urllib.parse import quote
            
            file_path = instance.archivo.path
            if os.path.exists(file_path):
                # Obtener el tipo de contenido basado en la extensión
                content_type, _ = guess_type(file_path)
                if not content_type:
                    content_type = 'application/octet-stream'
                
                # Obtener el nombre del archivo original (sin la ruta)
                filename = os.path.basename(instance.archivo.name)
                
                # Codificar el nombre del archivo para caracteres especiales
                # Usar RFC 5987 encoding para nombres de archivo con caracteres especiales
                filename_encoded = quote(filename, safe='')
                
                response = FileResponse(
                    open(file_path, 'rb'),
                    content_type=content_type
                )
                # Usar ambos formatos para máxima compatibilidad
                response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename_encoded}'
                return response
        
        # Comportamiento normal: devolver el serializer
        return super().retrieve(request, *args, **kwargs)


class InscripcionViewSet(viewsets.ModelViewSet):
    queryset = Inscripcion.objects.select_related('alumno', 'promocion', 'promocion__curso').all()
    serializer_class = InscripcionSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        
        # Alumnos solo ven sus propias inscripciones
        if user.es_alumno:
            queryset = queryset.filter(alumno=user)
        
        # Filtrar por promoción si se proporciona
        promocion_id = self.request.query_params.get('promocion')
        if promocion_id:
            queryset = queryset.filter(promocion_id=promocion_id)
        
        return queryset


class AsistenciaViewSet(viewsets.ModelViewSet):
    queryset = Asistencia.objects.select_related('inscripcion', 'inscripcion__alumno', 'tema').all()
    serializer_class = AsistenciaSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        
        # Alumnos solo ven sus propias asistencias
        if user.es_alumno:
            queryset = queryset.filter(inscripcion__alumno=user)
        
        # Filtrar por tema si se proporciona
        tema_id = self.request.query_params.get('tema')
        if tema_id:
            queryset = queryset.filter(tema_id=tema_id)
        
        return queryset


class PreguntaViewSet(viewsets.ModelViewSet):
    queryset = Pregunta.objects.select_related('tema').all()
    serializer_class = PreguntaSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        tema_id = self.request.query_params.get('tema')
        if tema_id:
            return self.queryset.filter(tema_id=tema_id)
        return self.queryset


class ExamenViewSet(viewsets.ModelViewSet):
    queryset = Examen.objects.select_related('tema', 'tema__curso').all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'list':
            return ExamenListSerializer
        return ExamenSerializer
    
    def get_queryset(self):
        from django.utils import timezone
        
        user = self.request.user
        queryset = super().get_queryset()
        
        # Alumnos solo ven exámenes de temas de cursos donde están inscritos
        if user.es_alumno:
            inscripciones = Inscripcion.objects.filter(
                alumno=user, activa=True
            ).select_related('promocion__curso')
            curso_ids = [insc.promocion.curso_id for insc in inscripciones]
            queryset = queryset.filter(
                tema__curso_id__in=curso_ids,
                tema__visible_para_estudiante=True,
                activo=True,
            )
            
            # Si no se está filtrando por tema específico, solo mostrar exámenes disponibles ahora
            tema_id = self.request.query_params.get('tema')
            if not tema_id:
                ahora = timezone.now()
                # Filtrar exámenes que estén disponibles en este momento
                # Debe estar activo Y (no tener fecha_inicio O fecha_inicio <= ahora) Y (no tener fecha_fin O fecha_fin >= ahora)
                queryset = queryset.filter(
                    activo=True
                ).filter(
                    Q(fecha_inicio__isnull=True) | Q(fecha_inicio__lte=ahora)
                ).filter(
                    Q(fecha_fin__isnull=True) | Q(fecha_fin__gte=ahora)
                )
        
        # Filtrar por tema si se proporciona
        tema_id = self.request.query_params.get('tema')
        if tema_id:
            queryset = queryset.filter(tema_id=tema_id)
        
        return queryset

    def get_object(self):
        """Permite acceder al examen vía recuperación aunque el examen original haya expirado."""
        user = self.request.user
        if not user.es_alumno:
            return super().get_object()

        recuperacion_id = self.request.query_params.get('recuperacion_id')
        if not recuperacion_id and self.request.method == 'POST':
            recuperacion_id = self.request.data.get('recuperacion_id')

        if recuperacion_id and self.action in ('retrieve', 'preguntas', 'responder'):
            from django.utils import timezone
            from rest_framework.exceptions import NotFound

            examen = get_object_or_404(
                Examen.objects.select_related('tema', 'tema__curso'),
                pk=self.kwargs['pk'],
            )
            inscripcion = Inscripcion.objects.filter(
                alumno=user,
                promocion__curso=examen.tema.curso,
                activa=True,
            ).first()
            if not inscripcion:
                raise NotFound('No estás inscrito en una promoción de este curso')

            ahora = timezone.now()
            recuperacion = RecuperacionExamen.objects.filter(
                id=recuperacion_id,
                examen=examen,
                inscripcion=inscripcion,
                activa=True,
                completada=False,
            ).filter(
                Q(fecha_inicio__isnull=True) | Q(fecha_inicio__lte=ahora)
            ).filter(
                Q(fecha_fin__isnull=True) | Q(fecha_fin__gte=ahora)
            ).first()

            if not recuperacion:
                raise NotFound('Recuperación no válida o no disponible')

            return examen

        return super().get_object()
    
    @action(detail=True, methods=['get'])
    def preguntas(self, request, pk=None):
        """Endpoint para obtener las preguntas aleatorias del examen para un estudiante"""
        examen = self.get_object()
        user = request.user
        
        if not user.es_alumno:
            return Response(
                {'error': 'Solo los alumnos pueden ver las preguntas del examen'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Verificar inscripción
        inscripcion = Inscripcion.objects.filter(
            alumno=user,
            promocion__curso=examen.tema.curso,
            activa=True
        ).first()
        
        if not inscripcion:
            return Response(
                {'error': 'No estás inscrito en una promoción de este curso'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        from django.utils import timezone
        ahora = timezone.now()
        
        # Verificar si hay una recuperación activa
        recuperacion_id = request.query_params.get('recuperacion_id')
        recuperacion = None
        if recuperacion_id:
            recuperacion = RecuperacionExamen.objects.filter(
                id=recuperacion_id,
                examen=examen,
                inscripcion=inscripcion,
                activa=True,
                completada=False
            ).first()
            
            if not recuperacion:
                return Response(
                    {'error': 'Recuperación no válida o no disponible'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Verificar fechas de recuperación
            if recuperacion.fecha_inicio and ahora < recuperacion.fecha_inicio:
                return Response(
                    {'error': 'La recuperación aún no está disponible'},
                    status=status.HTTP_403_FORBIDDEN
                )
            if recuperacion.fecha_fin and ahora > recuperacion.fecha_fin:
                return Response(
                    {'error': 'La recuperación ya expiró'},
                    status=status.HTTP_403_FORBIDDEN
                )
        else:
            # Examen normal - verificar fechas del examen
            if examen.fecha_inicio and ahora < examen.fecha_inicio:
                return Response(
                    {'error': 'El examen aún no está disponible'},
                    status=status.HTTP_403_FORBIDDEN
                )
            if examen.fecha_fin and ahora > examen.fecha_fin:
                return Response(
                    {'error': 'El examen ya expiró'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Verificar si ya tiene una calificación normal (sin recuperación)
            calificacion_normal = CalificacionExamen.objects.filter(
                examen=examen,
                inscripcion=inscripcion,
                recuperacion__isnull=True
            ).first()
            
            if calificacion_normal:
                return Response(
                    {'error': 'Ya has respondido este examen. Busca una recuperación si está disponible.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        # Obtener preguntas aleatorias
        preguntas = examen.obtener_preguntas_aleatorias()
        
        if preguntas.count() < examen.numero_preguntas:
            return Response(
                {'error': f'No hay suficientes preguntas en el banco. Se requieren al menos {examen.numero_preguntas} preguntas'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = PreguntaDetailSerializer(preguntas, many=True)
        
        return Response({
            'examen_id': examen.id,
            'recuperacion_id': recuperacion.id if recuperacion else None,
            'preguntas': serializer.data,
            'numero_preguntas': examen.numero_preguntas,
            'puntos_por_pregunta': examen.puntos_por_pregunta,
            'puntaje_total': examen.puntaje_total,
            'tiempo_limite': examen.tiempo_limite,
        })
    
    @action(detail=True, methods=['post'])
    def responder(self, request, pk=None):
        """Endpoint para que un alumno responda un examen (normal o recuperación)"""
        examen = self.get_object()
        user = request.user
        recuperacion_id = request.data.get('recuperacion_id')  # Opcional, para recuperaciones
        
        if not user.es_alumno:
            return Response(
                {'error': 'Solo los alumnos pueden responder exámenes'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Verificar inscripción en una promoción del mismo curso
        inscripcion = Inscripcion.objects.filter(
            alumno=user,
            promocion__curso=examen.tema.curso,
            activa=True
        ).first()
        
        if not inscripcion:
            return Response(
                {'error': 'No estás inscrito en una promoción de este curso'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        recuperacion = None
        if recuperacion_id:
            # Es una recuperación
            recuperacion = get_object_or_404(RecuperacionExamen, id=recuperacion_id, examen=examen, inscripcion=inscripcion)
            
            # Verificar que la recuperación esté activa y dentro de las fechas
            from django.utils import timezone
            ahora = timezone.now()
            if not recuperacion.activa:
                return Response({'error': 'Esta recuperación no está activa'}, status=status.HTTP_403_FORBIDDEN)
            if recuperacion.fecha_inicio and ahora < recuperacion.fecha_inicio:
                return Response({'error': 'La recuperación aún no está disponible'}, status=status.HTTP_403_FORBIDDEN)
            if recuperacion.fecha_fin and ahora > recuperacion.fecha_fin:
                return Response({'error': 'La recuperación ya expiró'}, status=status.HTTP_403_FORBIDDEN)
            
            # Verificar que no haya respondido ya esta recuperación
            calificacion_recuperacion = CalificacionExamen.objects.filter(
                examen=examen,
                inscripcion=inscripcion,
                recuperacion=recuperacion
            ).first()
            
            if calificacion_recuperacion:
                return Response(
                    {'error': 'Ya has respondido esta recuperación'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            # Es un examen normal
            # Verificar que no haya respondido ya el examen normal
            calificacion_normal = CalificacionExamen.objects.filter(
                examen=examen,
                inscripcion=inscripcion,
                recuperacion__isnull=True
            ).first()
            
            if calificacion_normal:
                return Response(
                    {'error': 'Ya has respondido este examen'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Procesar respuestas
        respuestas_data = request.data.get('respuestas', [])
        
        if len(respuestas_data) != examen.numero_preguntas:
            return Response(
                {'error': f'Debes responder exactamente {examen.numero_preguntas} preguntas'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        puntos_totales_obtenidos = 0
        for respuesta_data in respuestas_data:
            pregunta_id = respuesta_data.get('pregunta_id')
            respuesta_dada = respuesta_data.get('respuesta', '')
            
            pregunta = get_object_or_404(Pregunta, id=pregunta_id, tema=examen.tema)
            
            # Verificar si la respuesta es correcta (null-safe, case-insensitive)
            es_correcta = pregunta.es_respuesta_correcta(respuesta_dada)
            puntos_obtenidos = Decimal(examen.puntos_por_pregunta) if es_correcta else Decimal(0)
            puntos_totales_obtenidos += float(puntos_obtenidos)
            
            # Crear respuesta (con recuperación si aplica)
            RespuestaExamen.objects.create(
                examen=examen,
                inscripcion=inscripcion,
                pregunta=pregunta,
                recuperacion=recuperacion,
                respuesta_dada=respuesta_dada,
                es_correcta=es_correcta,
                puntos_obtenidos=puntos_obtenidos,
            )
        
        # Calcular calificación final
        calificacion = CalificacionExamen.objects.create(
            examen=examen,
            inscripcion=inscripcion,
            recuperacion=recuperacion,
            puntaje_obtenido=Decimal(puntos_totales_obtenidos),
            puntaje_total=Decimal(examen.puntaje_total),
            porcentaje=Decimal((puntos_totales_obtenidos / examen.puntaje_total) * 100) if examen.puntaje_total > 0 else Decimal(0),
        )
        
        # Si es recuperación, marcarla como completada
        if recuperacion:
            recuperacion.completada = True
            recuperacion.save()

        promedio, _ = PromedioPromocion.objects.get_or_create(inscripcion=inscripcion)
        promedio.calcular_promedio()
        
        serializer = CalificacionExamenSerializer(calificacion)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class RecuperacionExamenViewSet(viewsets.ModelViewSet):
    queryset = RecuperacionExamen.objects.select_related(
        'examen',
        'examen__tema',
        'examen__tema__curso',
        'inscripcion',
        'inscripcion__alumno',
    ).all()
    serializer_class = RecuperacionExamenSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        
        # Filtrar por examen si se proporciona
        examen_id = self.request.query_params.get('examen')
        if examen_id:
            queryset = queryset.filter(examen_id=examen_id)
        
        # Filtrar por inscripción si se proporciona
        inscripcion_id = self.request.query_params.get('inscripcion')
        if inscripcion_id:
            queryset = queryset.filter(inscripcion_id=inscripcion_id)
        
        # Alumnos solo ven sus propias recuperaciones
        if user.es_alumno:
            queryset = queryset.filter(inscripcion__alumno=user)
        
        return queryset
    
    def get_serializer_class(self):
        """Usar el serializer de creación múltiple si se envía una lista de inscripciones"""
        if self.action == 'create' and 'inscripciones' in self.request.data:
            # Si viene una lista de inscripciones, usar el serializer de creación múltiple
            if isinstance(self.request.data.get('inscripciones'), list):
                return RecuperacionExamenBulkCreateSerializer
        return RecuperacionExamenSerializer
    
    def create(self, request, *args, **kwargs):
        """Crear recuperación(es) - soporta creación individual o múltiple"""
        user = request.user
        if not (user.es_docente or user.is_superuser):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Solo los docentes pueden crear recuperaciones')
        
        # Detectar si es creación múltiple
        if 'inscripciones' in request.data and isinstance(request.data.get('inscripciones'), list):
            # Creación múltiple
            serializer = RecuperacionExamenBulkCreateSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            recuperaciones = serializer.save()
            omitidos = getattr(serializer, '_omitidos', [])
            
            # Retornar las recuperaciones creadas
            response_serializer = RecuperacionExamenSerializer(recuperaciones, many=True)
            response_data = {
                'recuperaciones': response_serializer.data,
                'creadas': len(recuperaciones),
            }
            if omitidos:
                response_data['omitidos'] = omitidos
                response_data['advertencia'] = (
                    f"No se asignó recuperación a quienes ya aprobaron: {', '.join(omitidos)}"
                )
            return Response(response_data, status=status.HTTP_201_CREATED)
        else:
            # Creación individual (compatibilidad hacia atrás)
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
    
    def perform_create(self, serializer):
        """Al crear una recuperación individual, validar que el usuario sea docente/admin"""
        user = self.request.user
        if not (user.es_docente or user.is_superuser):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Solo los docentes pueden crear recuperaciones')
        serializer.save()
    
    @action(detail=False, methods=['get'], url_path='mis-disponibles')
    def mis_disponibles(self, request):
        """Recuperaciones activas asignadas al alumno actual."""
        user = request.user
        if not user.es_alumno:
            return Response([])

        from django.utils import timezone
        from .serializers import RecuperacionExamenBulkCreateSerializer as BulkSerializer

        ahora = timezone.now()
        tema_id = request.query_params.get('tema')

        queryset = RecuperacionExamen.objects.filter(
            inscripcion__alumno=user,
            inscripcion__activa=True,
            activa=True,
            completada=False,
            examen__activo=True,
        ).filter(
            Q(fecha_inicio__isnull=True) | Q(fecha_inicio__lte=ahora)
        ).filter(
            Q(fecha_fin__isnull=True) | Q(fecha_fin__gte=ahora)
        ).select_related(
            'examen',
            'examen__tema',
            'examen__tema__curso',
            'inscripcion',
        )

        if tema_id:
            queryset = queryset.filter(examen__tema_id=tema_id)

        # Solo recuperaciones para quienes no aprobaron (original ni recuperación)
        elegibles = [
            rec for rec in queryset
            if BulkSerializer._inscripcion_puede_recuperacion(rec.examen, rec.inscripcion)
        ]

        serializer = RecuperacionExamenAlumnoSerializer(elegibles, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def contar_por_inscripcion(self, request):
        """Cuenta las recuperaciones totales de una inscripción (para límites)"""
        inscripcion_id = request.query_params.get('inscripcion_id')
        if not inscripcion_id:
            return Response(
                {'error': 'inscripcion_id es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        inscripcion = get_object_or_404(Inscripcion, id=inscripcion_id)
        total_recuperaciones = RecuperacionExamen.objects.filter(
            inscripcion=inscripcion,
            examen__tema__curso=inscripcion.promocion.curso
        ).count()
        
        return Response({
            'inscripcion_id': inscripcion_id,
            'total_recuperaciones': total_recuperaciones
        })


class CalificacionExamenViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CalificacionExamen.objects.select_related(
        'examen', 'examen__tema', 'inscripcion', 'inscripcion__alumno',
        'inscripcion__promocion', 'recuperacion'
    ).all()
    serializer_class = CalificacionExamenSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        
        # Alumnos solo ven sus propias calificaciones
        if user.es_alumno:
            queryset = queryset.filter(inscripcion__alumno=user)
        
        # Filtrar por examen si se proporciona
        examen_id = self.request.query_params.get('examen')
        if examen_id:
            queryset = queryset.filter(examen_id=examen_id)
        
        return queryset

    @action(detail=True, methods=['get'])
    def detalle(self, request, pk=None):
        """Detalle de calificación con respuestas (docentes/admins o el propio alumno)."""
        calificacion = self.get_object()
        user = request.user

        if user.es_alumno and calificacion.inscripcion.alumno_id != user.id:
            return Response(
                {'error': 'No tienes permiso para ver este examen'},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = CalificacionExamenDetalleSerializer(calificacion)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def revisar(self, request, pk=None):
        """Muestra las respuestas incorrectas del alumno una vez cerrado el examen."""
        calificacion = self.get_object()
        user = request.user

        if user.es_alumno and calificacion.inscripcion.alumno_id != user.id:
            return Response(
                {'error': 'No tienes permiso para revisar esta calificación'},
                status=status.HTTP_403_FORBIDDEN
            )

        if not calificacion.puede_revisar_respuestas():
            return Response(
                {'error': 'Las respuestas estarán disponibles cuando el examen se cierre por completo'},
                status=status.HTTP_403_FORBIDDEN
            )

        if calificacion.recuperacion:
            respuestas_qs = RespuestaExamen.objects.filter(
                examen=calificacion.examen,
                inscripcion=calificacion.inscripcion,
                recuperacion=calificacion.recuperacion,
            ).select_related('pregunta')
        else:
            respuestas_qs = RespuestaExamen.objects.filter(
                examen=calificacion.examen,
                inscripcion=calificacion.inscripcion,
                recuperacion__isnull=True,
            ).select_related('pregunta')

        # Si la clave se corrigió después del intento, sanear banderas obsoletas
        # (evita mostrar la misma opción como errónea y correcta a la vez).
        puntos_pregunta = Decimal(calificacion.examen.puntos_por_pregunta)
        hubo_correccion = False
        for respuesta in respuestas_qs:
            coincide_ahora = respuesta.pregunta.es_respuesta_correcta(respuesta.respuesta_dada)
            if coincide_ahora and not respuesta.es_correcta:
                respuesta.es_correcta = True
                respuesta.puntos_obtenidos = puntos_pregunta
                respuesta.save(update_fields=['es_correcta', 'puntos_obtenidos'])
                hubo_correccion = True

        if hubo_correccion:
            calificacion.calcular_calificacion()
            promedio, _ = PromedioPromocion.objects.get_or_create(
                inscripcion=calificacion.inscripcion
            )
            promedio.calcular_promedio()

        respuestas_incorrectas = [
            r for r in respuestas_qs
            if not r.es_correcta and not r.pregunta.es_respuesta_correcta(r.respuesta_dada)
        ]

        serializer = RespuestaRevisionSerializer(respuestas_incorrectas, many=True)
        return Response({
            'calificacion_id': calificacion.id,
            'examen_titulo': calificacion.examen.titulo,
            'tema_titulo': calificacion.examen.tema.titulo,
            'total_incorrectas': len(respuestas_incorrectas),
            'respuestas_incorrectas': serializer.data,
        })


class PromedioPromocionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = PromedioPromocion.objects.select_related('inscripcion', 'inscripcion__alumno', 'inscripcion__promocion').all()
    serializer_class = PromedioPromocionSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        
        # Alumnos solo ven sus propios promedios
        if user.es_alumno:
            queryset = queryset.filter(inscripcion__alumno=user)
        
        # Filtrar por promoción si se proporciona
        promocion_id = self.request.query_params.get('promocion')
        if promocion_id:
            queryset = queryset.filter(inscripcion__promocion_id=promocion_id)
        
        return queryset
    
    @action(detail=False, methods=['post'])
    def calcular_promedios(self, request):
        """Endpoint para calcular promedios de una promoción"""
        promocion_id = request.data.get('promocion_id')
        if not promocion_id:
            return Response(
                {'error': 'promocion_id es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        from .models import Inscripcion
        inscripciones = Inscripcion.objects.filter(promocion_id=promocion_id, activa=True)
        
        for inscripcion in inscripciones:
            promedio, created = PromedioPromocion.objects.get_or_create(
                inscripcion=inscripcion
            )
            promedio.calcular_promedio()
        
        return Response({'mensaje': 'Promedios calculados correctamente'})


def _normalize_course_name(name):
    if not name:
        return ''
    normalized = name.strip().lower()
    if normalized.startswith('-'):
        normalized = normalized.lstrip('-').strip()
    return normalized


def _get_brittany_font(font_size):
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ModuleNotFoundError:
        return None
    font_paths = [
        os.path.join(settings.BASE_DIR, 'cursos', 'assets', 'fonts', 'Brittany.ttf'),
        os.path.join(settings.MEDIA_ROOT, 'fonts', 'Brittany.ttf'),
    ]
    for font_path in font_paths:
        if os.path.exists(font_path):
            font_name = 'Brittany'
            if font_name not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont(font_name, font_path))
            return font_name
    return None


def _generate_diploma_pdf(alumno_nombre, curso_nombre):
    try:
        from PyPDF2 import PdfReader, PdfWriter
        from reportlab.pdfgen import canvas
        from reportlab.lib.colors import HexColor
        from reportlab.pdfbase import pdfmetrics
    except ModuleNotFoundError:
        return None, 'Faltan dependencias para generar PDFs (reportlab y PyPDF2).'
    template_map = {
        'escuela de corderitos': {
            'template_filename': 'Escuela de corderitos diploma.pdf',
            'text_color': '#1f3b5a',
            'font_size': 48,
            'y_ratio': 0.54,
            'max_width_ratio': 0.75,
        },
        'escuela de doctrina intermedia': {
            'template_filename': 'Diplomas Doctrina Intermedia.pdf',
            'text_color': '#1f3b5a',
            'font_size': 48,
            'y_ratio': 0.54,
            'max_width_ratio': 0.75,
        },
    }
    normalized_course = _normalize_course_name(curso_nombre)
    if normalized_course not in template_map:
        return None, 'No hay plantilla configurada para este curso.'

    config = template_map[normalized_course]
    template_path = os.path.join(
        settings.BASE_DIR.parent,
        'frontend',
        'public',
        'images',
        config['template_filename'],
    )
    if not os.path.exists(template_path):
        return None, 'No se encontró la plantilla del diploma.'

    reader = PdfReader(template_path)
    if not reader.pages:
        return None, 'La plantilla del diploma está vacía.'

    base_page = reader.pages[0]
    width = float(base_page.mediabox.width)
    height = float(base_page.mediabox.height)

    font_size = config['font_size']
    font_name = _get_brittany_font(font_size) or 'Helvetica'
    max_text_width = width * config['max_width_ratio']
    text_width = pdfmetrics.stringWidth(alumno_nombre, font_name, font_size)
    while text_width > max_text_width and font_size > 28:
        font_size -= 2
        text_width = pdfmetrics.stringWidth(alumno_nombre, font_name, font_size)

    x = (width - text_width) / 2
    y = height * config['y_ratio']

    packet = io.BytesIO()
    overlay = canvas.Canvas(packet, pagesize=(width, height))
    overlay.setFillColor(HexColor(config['text_color']))
    overlay.setFont(font_name, font_size)
    overlay.drawString(x, y, alumno_nombre)
    overlay.save()

    packet.seek(0)
    overlay_reader = PdfReader(packet)
    base_page.merge_page(overlay_reader.pages[0])

    writer = PdfWriter()
    writer.add_page(base_page)

    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output, None


class DiplomaViewSet(viewsets.ModelViewSet):
    queryset = Diploma.objects.select_related('inscripcion', 'inscripcion__alumno', 'inscripcion__promocion').all()
    serializer_class = DiplomaSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        
        # Alumnos solo ven sus propios diplomas
        if user.es_alumno:
            queryset = queryset.filter(inscripcion__alumno=user)

        promocion_id = self.request.query_params.get('promocion')
        if promocion_id:
            queryset = queryset.filter(inscripcion__promocion_id=promocion_id)
        
        return queryset

    def retrieve(self, request, *args, **kwargs):
        """Sobrescribir retrieve para servir el archivo del diploma"""
        instance = self.get_object()

        if request.query_params.get('download') == 'true' and instance.archivo:
            from django.http import FileResponse
            from mimetypes import guess_type
            from urllib.parse import quote

            file_path = instance.archivo.path
            if os.path.exists(file_path):
                content_type, _ = guess_type(file_path)
                if not content_type:
                    content_type = 'application/octet-stream'

                filename = os.path.basename(instance.archivo.name)
                filename_encoded = quote(filename, safe='')

                response = FileResponse(open(file_path, 'rb'), content_type=content_type)
                response['Content-Disposition'] = (
                    f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename_encoded}'
                )
                return response

        return super().retrieve(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def descargar_zip(self, request):
        """Descargar diplomas de una promoción en un ZIP"""
        promocion_id = request.query_params.get('promocion_id') or request.query_params.get('promocion')
        if not promocion_id:
            return Response({'error': 'promocion_id es requerido'}, status=status.HTTP_400_BAD_REQUEST)

        promocion = get_object_or_404(Promocion, id=promocion_id)
        diplomas = (
            self.get_queryset()
            .filter(inscripcion__promocion_id=promocion_id)
            .select_related('inscripcion__alumno', 'inscripcion__promocion__curso')
        )

        zip_buffer = io.BytesIO()
        added = 0
        import zipfile
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for diploma in diplomas:
                if diploma.archivo and os.path.exists(diploma.archivo.path):
                    filename = os.path.basename(diploma.archivo.name)
                    zip_file.write(diploma.archivo.path, arcname=filename)
                    added += 1

        if added == 0:
            return Response(
                {'error': 'No hay diplomas disponibles para descargar'},
                status=status.HTTP_404_NOT_FOUND
            )

        zip_buffer.seek(0)
        from django.http import HttpResponse
        from urllib.parse import quote

        base_name = promocion.nombre or f'promocion_{promocion_id}'
        safe_name = ''.join(ch if ch.isalnum() or ch in ('-', '_') else '_' for ch in base_name)
        safe_name = safe_name.strip('_') or f'promocion_{promocion_id}'
        zip_filename = f'Diplomas_{safe_name}.zip'
        zip_filename_encoded = quote(zip_filename, safe='')

        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = (
            f'attachment; filename="{zip_filename}"; filename*=UTF-8\'\'{zip_filename_encoded}'
        )
        return response
    
    @action(detail=False, methods=['post'])
    def generar_diplomas(self, request):
        """Endpoint para generar diplomas para estudiantes aprobados de una promoción"""
        promocion_id = request.data.get('promocion_id')
        if not promocion_id:
            return Response(
                {'error': 'promocion_id es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            promocion = get_object_or_404(Promocion, id=promocion_id)
            total_examenes = Examen.objects.filter(tema__curso=promocion.curso).count()
            if total_examenes == 0:
                return Response({
                    'mensaje': 'No hay exámenes configurados para este curso.',
                    'diplomas': [],
                    'advertencias': [],
                })

            # Recalcular promedios antes de evaluar aprobados
            inscripciones = Inscripcion.objects.filter(promocion=promocion, activa=True)
            for inscripcion in inscripciones:
                promedio, _created = PromedioPromocion.objects.get_or_create(
                    inscripcion=inscripcion
                )
                promedio.calcular_promedio()

            # Obtener inscripciones con promedio aprobado (>= 80%)
            promedios = PromedioPromocion.objects.filter(
                inscripcion__promocion_id=promocion_id,
                aprobado=True
            )

            completados_por_inscripcion = {
                item['inscripcion_id']: item['examenes_completados']
                for item in CalificacionExamen.objects.filter(
                    inscripcion__promocion_id=promocion_id,
                    examen__tema__curso=promocion.curso
                )
                .values('inscripcion_id')
                .annotate(examenes_completados=Count('examen', distinct=True))
            }
            
            diplomas_creados = []
            diplomas_resultados = []
            advertencias = []
            for promedio in promedios:
                completados = completados_por_inscripcion.get(promedio.inscripcion_id, 0)
                if completados < total_examenes:
                    alumno_nombre = (
                        promedio.inscripcion.alumno.get_full_name()
                        or promedio.inscripcion.alumno.username
                    )
                    advertencias.append({
                        'alumno': alumno_nombre,
                        'curso': promedio.inscripcion.promocion.curso.nombre,
                        'detalle': (
                            f'No ha completado todos los exámenes '
                            f'({completados}/{total_examenes}).'
                        ),
                    })
                    continue

                # Verificar si ya tiene diploma
                diploma, created = Diploma.objects.get_or_create(
                    inscripcion=promedio.inscripcion,
                    defaults={'activo': True}
                )
                alumno_nombre = (
                    promedio.inscripcion.alumno.get_full_name()
                    or promedio.inscripcion.alumno.username
                )
                curso_nombre = promedio.inscripcion.promocion.curso.nombre
                if created or not diploma.archivo:
                    pdf_buffer, error = _generate_diploma_pdf(alumno_nombre, curso_nombre)
                    if pdf_buffer:
                        filename = f"diploma_{diploma.codigo_diploma}.pdf"
                        diploma.archivo.save(filename, ContentFile(pdf_buffer.read()), save=True)
                    elif error:
                        advertencias.append({
                            'alumno': alumno_nombre,
                            'curso': curso_nombre,
                            'detalle': error,
                        })
                if created:
                    diplomas_creados.append({
                        'alumno': alumno_nombre,
                        'codigo': diploma.codigo_diploma,
                        'archivo': diploma.archivo.url if diploma.archivo else None,
                        'creado': True,
                    })
                diplomas_resultados.append({
                    'alumno': alumno_nombre,
                    'codigo': diploma.codigo_diploma,
                    'archivo': diploma.archivo.url if diploma.archivo else None,
                    'creado': created,
                })
            
            return Response({
                'mensaje': f'Diplomas generados: {len(diplomas_creados)}',
                'diplomas': diplomas_resultados,
                'advertencias': advertencias,
            })
        except Exception as exc:
            return Response(
                {'error': f'Error al generar diplomas: {exc}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

